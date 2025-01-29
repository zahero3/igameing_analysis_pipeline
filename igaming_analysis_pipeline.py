import pandas as pd
import numpy as np
import random
import psycopg2
from airflow import DAG
from airflow.operators.python import PythonOperator
from datetime import datetime, timedelta


def strt_analysis(filePath="",fileName=""):
    
    # Data Extraction
    if filePath == "" or fileName == "":
        # creating connection
        conn = psycopg2.connect(
            host="localhost",
            database="postgres",
            user="home",
            password="233")
        
        #cursor = conn.cursor() # Cursor is used if we want to execute  (CREATE, INSERT, DELETE)
        
        query = "SELECT * FROM practice.igaming"
        df = pd.read_sql(query, conn)
    else:
        
        df = pd.read_csv(f'{filePath}/{fileName}') # If data is in csv file
    
    #print(df)
    trans(df)
    
def trans(data):

    df = data
    
    # Map engagement levels to in-game purchase spending ranges
    purchase_ranges = {
        "Low": (0, 5),     # Low engagement results in low spending
        "Medium": (5, 50), # Medium engagement results in moderate spending
        "High": (50, 200)  # High engagement results in high spending
    }
    
    # Generate random in-game purchase values based on engagement level
    df["InGamePurchases"] = df["EngagementLevel"].map(
        lambda x: round(random.uniform(*purchase_ranges[x]), 2)
    )
    
    # Add simulated Total_Bets and Total_Winnings
    # Use engagement metrics to estimate bets (scaling by 10 for realistic values)
    df["Total_Bets"] = (df["PlayTimeHours"] * df["SessionsPerWeek"] * df["AvgSessionDurationMinutes"]) / 10
    
    # Simulate RTP based on GameGenre (different games have different RTP ranges)
    # For simplicity, we use a random RTP between 70% and 95%.
    np.random.seed(42)  # Set seed for reproducibility
    df["RTP"] = np.random.uniform(0.7, 0.95, len(df))  # Random RTP for each row
    df["Total_Winnings"] = df["Total_Bets"] * df["RTP"]
    
    # Calculate GGR (Gross Gaming Revenue)
    df["GGR"] = df["Total_Bets"] - df["Total_Winnings"]
    
    # Calculate NGR (Net Gaming Revenue)
    df["NGR"] = df["GGR"] - df["InGamePurchases"]
    
    # Estimate Revenue for players without InGamePurchases based on Engagement
    engagement_factors = {"Low": 0.5, "Medium": 1, "High": 2}
    df["Engagement_Factor"] = df["EngagementLevel"].map(engagement_factors)
    #df["Estimated_Revenue"] = df["Engagement_Factor"] * df["SessionsPerWeek"] * df["AvgSessionDurationMinutes"]

    
    # Creating new excel file and saving the transformed data in it
    with pd.ExcelWriter("iGamingAnalysisOutput.xlsx", engine="openpyxl", mode="w") as writer:
        df.to_excel(writer, sheet_name="TransformedData", index=False)
        
    print("Transformation done on the dataset.")
    analyze(df)

def analyze(data):
    df = data
    trns_df = df.copy() # Transformed Data
    
    ########################### Data Pre Processing for analysis #########################################
    
    #print(df.isnull().sum()) # Check for missing values
    df = df.dropna() # Dropping missing values

    from scipy.stats import zscore
    df = df[(zscore(df["PlayTimeHours"]) < 3)] # Removing outliers from PlayTimeHours
    
    print("Data processing Done.")


    ########################### Player Segmentation ######################################################

    from sklearn.cluster import KMeans
    from sklearn.preprocessing import StandardScaler
    
    # Select features for clustering
    # Behavioural Features
    features = df[["PlayTimeHours", "SessionsPerWeek", "AvgSessionDurationMinutes", "InGamePurchases"]]
    
    # Normalize the data
    scaler = StandardScaler()
    scaled_features = scaler.fit_transform(features)
    
    # Apply K-Means clustering
    kmeans = KMeans(n_clusters=3, random_state=42)
    df["PlayerType"] = kmeans.fit_predict(scaled_features)

    segment_mapping = {0: "Casual Players", 1: "Regular Players", 2: "Hardcore Gamers"} # Define mapping for player segments
    
    #numeric_df = df.select_dtypes(include=['number']) # Select only numeric columns
    #print(numeric_df)
    cols_df = df[["PlayTimeHours", "SessionsPerWeek", "AvgSessionDurationMinutes", "InGamePurchases", "PlayerType"]]
    
    pf1 = df.groupby("PlayerType")[cols_df.columns].mean() # Group by "PlayerSegment" and calculate the mean of numeric columns
    pf1["PlayerType"] = pf1["PlayerType"].map(segment_mapping) # Replace numeric segments with meaningful labels

    # Player segments grouped
    with pd.ExcelWriter("iGamingAnalysisOutput.xlsx", engine="openpyxl", mode="a") as writer:
        pf1.to_excel(writer, sheet_name="PlayerType_Grouped", index=False) 

    pf2 = cols_df.copy()
    pf2["PlayerType"] = pf2["PlayerType"].map(segment_mapping)
    #pf2.insert(0, "PlayerID", df["PlayerID"])  # Add PlayerID column at the beginning

    trns_df["PlayerType"] = pf2["PlayerType"].copy() # Adding mapped player segment to Transformed Data

    with pd.ExcelWriter("iGamingAnalysisOutput.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        trns_df.to_excel(writer, sheet_name="TransformedData", index=False)

        ######################## Other Segment ########################
    # Revenue Features
    features2 = df[["Total_Bets", "Total_Winnings", "GGR", "NGR"]]
    
    # Normalize the data
    scaler2 = StandardScaler()
    scaled_features2 = scaler2.fit_transform(features2)
    
    # Apply K-Means clustering
    kmeans2 = KMeans(n_clusters=3, random_state=42)
    df["SpenderType"] = kmeans2.fit_predict(scaled_features2)

    segment_mapping2 = {0: "Low Spender", 1: "Moderate Spender", 2: "High Spender"} # Define mapping for player segments
    
    #numeric_df = df.select_dtypes(include=['number']) # Select only numeric columns
    #print(numeric_df)
    cols_df2 = df[["Total_Bets", "Total_Winnings", "GGR", "NGR", "SpenderType"]]
    
    pf1 = df.groupby("SpenderType")[cols_df2.columns].mean() # Group by "PlayerSegment" and calculate the mean of numeric columns
    pf1["SpenderType"] = pf1["SpenderType"].map(segment_mapping2) # Replace numeric segments with meaningful labels

    # Player segments grouped
    with pd.ExcelWriter("iGamingAnalysisOutput.xlsx", engine="openpyxl", mode="a") as writer:
        pf1.to_excel(writer, sheet_name="SpenderType_Grouped", index=False) 

    pf2 = cols_df2.copy()
    pf2["SpenderType"] = pf2["SpenderType"].map(segment_mapping2)
    #pf2.insert(0, "PlayerID", df["PlayerID"])  # Add PlayerID column at the beginning

    trns_df["SpenderType"] = pf2["SpenderType"].copy() # Adding mapped player segment to Transformed Data

    with pd.ExcelWriter("iGamingAnalysisOutput.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        trns_df.to_excel(writer, sheet_name="TransformedData", index=False)
    
    print("Player Segmentation Done.")

    
    ########################### Revenue Analysis ###################################################### 

    total_bets = round(df["Total_Bets"].sum(),2)
    total_wins = round(df["Total_Winnings"].sum(),2)

    # Total Revenue
    total_revenue = round((total_bets - total_wins) + df["InGamePurchases"].sum(),2) 
    print(f"Total Revenue: {total_revenue}")

    # GGR
    ggr = round((total_bets - total_wins),2)
    print(f"GGR: {ggr}")
    
    def calculate_bonus(row):
        if row["SpenderType"] == "Low Spender":
            return round(((20 * row["GGR"]) / 100), 2)  # 20% Bonus for Low Spender
        elif row["SpenderType"] == "Moderate Spender":
            return round(((30 * row["GGR"]) / 100), 2)  # 30% Bonus for Moderate Spender
        else:
            return round(((40 * row["GGR"]) / 100), 2)  # 40% Bonus for High Spender
    
    df["Bonus"] = df.apply(calculate_bonus, axis=1)
    trns_df["Bonus"] = df["Bonus"].copy() # Adding Bonus col to Transformed Data
    with pd.ExcelWriter("iGamingAnalysisOutput.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        trns_df.to_excel(writer, sheet_name="TransformedData", index=False)
    
    bns = round(df["Bonus"].sum(),2)
    
    # NGR
    ngr = round((ggr - bns),2) 
    print(f"NGR: {ngr}")

    rtp = round(((total_wins/total_bets) * 100),2)
    print(f"RTP: {rtp}%")

    arpu = round(total_revenue / df["PlayerID"].nunique(), 2) # ARPU
    print(f"ARPU: {arpu}")

    btw = round((total_bets / total_wins),2)
    print(f"Bet to Win ratio: {btw}")

    bcr = round(((bns / total_bets) * 100),2)
    print(f"Bonus Cost percentage: {bcr}%")

    hf = df[df["InGamePurchases"] > df["InGamePurchases"].quantile(0.9)] # Top 10% spenders

    # HighSpenders Saved to Excel sheet
    with pd.ExcelWriter("iGamingAnalysisOutput.xlsx", engine="openpyxl", mode="a") as writer:
        hf.to_excel(writer, sheet_name="Top10Prcnt_Spndr", index=False) 
        
    print("Revenue Analysis Done.")

    
    ########################### Retention Analysis #####################################################

    df["Average_time_in_mins"] = df["SessionsPerWeek"] * df["AvgSessionDurationMinutes"] # Creating time column in mins
    #df["Average_time_in_mins"].to_excel("time_in_mins.xlsx", index=False)

    # Retention Rate (Let it be defined as players playing more than 300 mins in a week are more likely to be retained)
    rr = len(df[df["Average_time_in_mins"] > 300])/len(df) 
    print(f'Retention rate: {rr * 100:.2f}%')
    r_prcnt = "{:.2f}%".format(rr * 100)

    print("Retention Analysis Done.")

    
    ########################### Engagement Analysis #####################################################
    
    # Most Engaging Game: Analyze which games have the highest playtime or engagement:
    eg = df.groupby("GameGenre")["PlayTimeHours"].sum().reset_index().sort_values(ascending = False, by="PlayTimeHours")
    
    with pd.ExcelWriter("iGamingAnalysisOutput.xlsx", engine="openpyxl", mode="a") as writer:
        eg.to_excel(writer, sheet_name="MostEngagingGames", index=False) # MostEngagingGames
        
    print("Engagement Analysis Done.")

    
    ########################### Prediction Models ###################################################### 
    
    from sklearn.model_selection import train_test_split
    from sklearn.ensemble import RandomForestClassifier
    from sklearn.ensemble import RandomForestRegressor
    
        
    ########################### Revenue Prediction Model - using RandomForestRegressor #####################
    def revenue_model(df):

        #Creating New Revenue Column for model
        df["Total_Revenue"] = (df["Total_Bets"]-df["Total_Winnings"]) + df["InGamePurchases"]
        
        # Define features and target
        features = df[["Total_Bets","Total_Winnings","PlayTimeHours", "SessionsPerWeek", 
                       "AvgSessionDurationMinutes", "InGamePurchases"]] # Independent variables/columns
        target = df["Total_Revenue"]  # Dependent variable/column
        
        # Train-test split
        X_train, X_test, y_train, y_test = train_test_split(features, target, test_size=0.2, random_state=42)
        
        # Train Random Forest Regressor
        rpm = RandomForestRegressor(random_state=42)
        rpm.fit(X_train, y_train)
        
        # Evaluate the model
        accuracy = rpm.score(X_test, y_test)
        print(f"Revenue Prediction Accuracy: {accuracy * 100:.2f}%")
        
        print("Model for Revenue Prediction created as rpm.")
        return rpm

    # Calling Trained model to get values for matrices
    estimated_revn = revenue_model(df).predict(df[["Total_Bets","Total_Winnings","PlayTimeHours", "SessionsPerWeek", 
                                    "AvgSessionDurationMinutes", "InGamePurchases"]]) # Returns predicted value for each row

    trns_df["Estimated_Revenue"] = estimated_revn # Pasting the estimated revenue column to transformed data sheet
    with pd.ExcelWriter("iGamingAnalysisOutput.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        trns_df.to_excel(writer, sheet_name="TransformedData", index=False)
        
    estimated_revn = round(estimated_revn.sum(),2) # Summing estimated revenue to show in Matrices
    
    ########################### Churn Prediction Model - using RandomForestClassifier ##################

    # Creating Churned column for training the model
    df['Churned'] = ((df['PlayTimeHours'] < 4) & (df['SessionsPerWeek'] < 14) & 
                     (df['AvgSessionDurationMinutes'] < 30) & 
                        (df['InGamePurchases'] < 5)).astype(int)

    """
    # Using lambda function
    df['Churned_lambda'] = df.apply(
        lambda row: 1 if (row['PlayTimeHours'] > 4 and 
                          row['SessionsPerWeek'] == 6 and 
                          row['AvgSessionDurationMinutes'] > 20 and 
                          row['InGamePurchases'] > 0) else 0,
        axis=1)
    """
    trns_df["Churns"] = df['Churned'].copy() # Pasting the churned column to transformed data sheet
    with pd.ExcelWriter("iGamingAnalysisOutput.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        trns_df.to_excel(writer, sheet_name="TransformedData", index=False)
            
    def churn_model(df):

        # Define features and target
        features = df[["PlayTimeHours", "SessionsPerWeek", "AvgSessionDurationMinutes", "InGamePurchases"]] # Independent variables/columns
        target = df["Churned"]  # Dependent variable/column
        
        # Train-test split
        X_train, X_test, y_train, y_test = train_test_split(features, target, test_size=0.2, random_state=42)
        
        # Train Random Forest Classifier
        cpm = RandomForestClassifier(random_state=42)
        cpm.fit(X_train, y_train)
        
        # Evaluate the model
        accuracy = cpm.score(X_test, y_test)
        print(f"Churn Prediction Accuracy: {accuracy * 100:.2f}%")
        
        print("Model for Churn Prediction created as cpm.")
        return cpm

    # Calling Trained model to get values for matrices
    cpm = churn_model(df).predict(df[["PlayTimeHours", "SessionsPerWeek", "AvgSessionDurationMinutes", "InGamePurchases"]])

    churnRate = round(((df["Churned"].sum() / df["Churned"].count()) * 100),2)
    print(f"Churn Rate: {churnRate}%")
    
    
    ########################### Matrices ###################################################### 

    from openpyxl import load_workbook
    
    matrix = pd.DataFrame({"Metrics": ["Total Revenue","Gross Gaming Revenue (GGR)","Net Gaming Revenue (NGR)","Return to Player (RTP)",
                                       "Average Revenue Per User (ARPU)", "Bet to Win Ratio","Bonus Cost Ratio",
                                       "Active Players","Retention Rate", "Most Engageing GameGenre","Bonus", 
                                       "Estimated Revenue", "Churn Rate"],
               "Values": [total_revenue, ggr,ngr, f'{rtp}%', arpu, btw, f'{bcr}%', df["PlayerID"].nunique(),r_prcnt, eg.iloc[0,0], bns,
                          estimated_revn, f'{churnRate}%']})

    # Saved to Excel
    with pd.ExcelWriter("iGamingAnalysisOutput.xlsx", engine="openpyxl", mode="a") as writer:
        matrix.to_excel(writer, sheet_name="SummaryMatrices", index=False) 

    # Moving Summary sheet at first position
    workbook = load_workbook("iGamingAnalysisOutput.xlsx") # Load the workbook    
    sheet_to_move = workbook["SummaryMatrices"] # Get the sheet you want to move
    workbook.move_sheet(sheet_to_move, offset=-workbook.index(sheet_to_move)) # Move the sheet to the first position
    workbook.save("iGamingAnalysisOutput.xlsx") # Save the workbook

    print("Matrices Calculated.")


    ########################### Visualization ###################################################### 
        
    import matplotlib.pyplot as plt
    import seaborn as sns
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.utils import get_column_letter
    
    # Load data
    df = pd.read_excel("iGamingAnalysisOutput.xlsx", sheet_name="TransformedData")
    
    # Create folder for visuals
    import os
    if not os.path.exists("visuals"):
        os.makedirs("visuals")
    
    # Load the existing workbook
    wb = load_workbook("iGamingAnalysisOutput.xlsx")
    
    # Check if "Visuals" sheet exists, if not create it
    if "Visuals" not in wb.sheetnames:
        ws = wb.create_sheet(title="Visuals")
    else:
        ws = wb["Visuals"]
    
    # Helper function to save plots and add them to Excel
    def save_and_add_plot(fig, name, row):
        image_path = f"visuals/{name}.png"
        fig.savefig(image_path, bbox_inches="tight")
        img = OpenpyxlImage(image_path)
        ws.add_image(img, f"A{row}")
        plt.close(fig)
    
    # Visual 1: Player Spending Segmentation Distribution
    fig, ax = plt.subplots(figsize=(8, 5))
    df['SpenderType'].value_counts().plot(kind='pie', autopct='%1.1f%%', startangle=90, ax=ax)
    ax.set_ylabel("")
    ax.set_title("Player Spending Distribution")
    save_and_add_plot(fig, "PlayerSegmentation", 1)
    """
    # Visual 1.1: Player Type Segmentation Distribution
    fig, ax = plt.subplots(figsize=(8, 5))
    df['PlayerType'].value_counts().plot(kind='pie', autopct='%1.1f%%', startangle=90, ax=ax)
    ax.set_ylabel("")
    ax.set_title("Player Type Segmentation Distribution")
    save_and_add_plot(fig, "PlayerSegmentation1", 1)
    """
    
    # Visual 2: Revenue Average Metrics by SpenderType
    fig, ax = plt.subplots(figsize=(10, 6))
    segment_means = df.groupby("SpenderType")[["Total_Bets", "Total_Winnings", "GGR", "NGR"]].mean()
    segment_means.plot(kind="bar", ax=ax)
    ax.set_title("Revenue Average Metrics by Spender Type")
    ax.set_ylabel("Average Value")
    save_and_add_plot(fig, "AvgMetricsBySpenderType", 20)

    """
    # Visual 2.1: Behavioural Average Metrics by PlayerType
    fig, ax = plt.subplots(figsize=(10, 6))
    segment_means = df.groupby("PlayerType")[["PlayTimeHours", "SessionsPerWeek", "AvgSessionDurationMinutes", "InGamePurchases"]].mean()
    segment_means.plot(kind="bar", ax=ax)
    ax.set_title("Behavioural Average Metrics by PlayerType")
    ax.set_ylabel("Average Value")
    save_and_add_plot(fig, "AvgMetricsByPlayerType", 20)
    """
    
    # Visual 3: Retention Rate Over Time (Example Time-Series Data)
    # Add fake time-series data for demonstration
    start_date = "2024-01-01"
    end_date = "2024-12-25"
    date_pool = pd.date_range(start = start_date, end = end_date)
    
    df["date"] = np.random.choice(date_pool, size= len(df), replace = True)
    retention_df = df.groupby("date")["PlayerID"].count()
    fig, ax = plt.subplots(figsize=(10, 5))
    retention_df.plot(ax=ax)
    ax.set_title("Retention Rate Over Time")
    ax.set_ylabel("Retention Count")
    save_and_add_plot(fig, "RetentionRate", 40)
    
    # Visual 4: Estimated Revenue vs Playtime
    fig, ax = plt.subplots(figsize=(8, 6))
    sns.scatterplot(data=df, x="PlayTimeHours", y="Estimated_Revenue", hue="SpenderType", ax=ax)
    ax.set_title("Estimated Revenue vs Playtime")
    save_and_add_plot(fig, "RevenueVsPlaytime", 60)
    
    # Visual 5: Bonus Cost Ratio by Player Segment
    df["Bonus_Cost_Ratio"] = df["Bonus"] / df["Total_Bets"]
    #df["Bonus_Cost_Ratio"] = (round(((30 * df["GGR"])/100),2) / df["Total_Bets"]) # 30% Bonus divide by total bets

    fig, ax = plt.subplots(figsize=(8, 5))
    bonus_means = df.groupby("SpenderType")["Bonus_Cost_Ratio"].mean()
    bonus_means.plot(kind="bar", color="skyblue", ax=ax)
    ax.set_title("Bonus Cost Ratio by Player Segment")
    save_and_add_plot(fig, "BonusCostRatio", 80)
    
    # Visual 6: Churned vs Retained Players
    fig, ax = plt.subplots(figsize=(8, 5))
    churned = df["Churns"].value_counts()
    churned.plot(kind="bar", color=["lightgreen", "salmon"], ax=ax)
    ax.set_title("Retained vs Churned Players")
    save_and_add_plot(fig, "ChurnedVsRetained", 100)
    
    # Visual 7: Player Engagement by Game Genre
    fig, ax = plt.subplots(figsize=(10, 6))
    engagement = df.groupby("GameGenre")[["PlayTimeHours", "SessionsPerWeek"]].mean()
    sns.heatmap(engagement, annot=True, cmap="coolwarm", ax=ax)
    ax.set_title("Player Engagement by Game Genre")
    save_and_add_plot(fig, "EngagementByGenre", 120)
    
    # Visual 8: Player Activity Trend
    fig, ax = plt.subplots(figsize=(10, 5))
    activity_trend = df.groupby("date")["PlayTimeHours"].mean()
    activity_trend.plot(ax=ax)
    ax.set_title("Player Activity Trend")
    save_and_add_plot(fig, "ActivityTrend", 140)
    
    # Visual 9: In-Game Purchases Distribution
    fig, ax = plt.subplots(figsize=(8, 5))
    sns.boxplot(data=df, x="SpenderType", y="InGamePurchases", ax=ax)
    ax.set_title("In-Game Purchases Distribution")
    save_and_add_plot(fig, "PurchasesDistribution", 160)
    
    # Visual 10: High Spender Behavior
    fig, ax = plt.subplots(figsize=(8, 5))
    high_spender = df[df["SpenderType"] == "High Spender"]
    sns.kdeplot(data=high_spender, x="GGR", y="NGR", cmap="Blues", ax=ax, fill=True)
    ax.set_title("High Spender Behavior")
    save_and_add_plot(fig, "HighSpenderBehavior", 180)
    
    # Save Workbook
    #wb.save("iGamingAnalysisOutput_with_Visuals.xlsx")
    #print("Visuals created and saved to 'iGamingAnalysisOutput_with_Visuals.xlsx'.")

    # Saved to Excel
    wb.save("iGamingAnalysisOutput.xlsx")
    print("Plots created.")



# Define Dag

dag = DAG(
    'igaming_analysis_pipeline',
    #schedule_interval='@daily',
    schedule_interval=timedelta(seconds=300),  # Runs every 300 seconds
    start_date=datetime(2023, 1, 1),
    catchup=False
)

task = PythonOperator(
    task_id='run_igaming_analysis_etl',
    python_callable=strt_analysis,
    dag=dag
)